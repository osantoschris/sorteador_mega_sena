from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.formula.translate import Translator
import os

# Identifica par ou ímpar
def par_impar(numero):
    resto = numero % 2
    if resto == 0:
        return True
    else:
        return False

# Abrir o arquivo
def abrir_arquivo():
    os.startfile('C:\\Gerador Mega Sena\\bases\\sorteios_contabilizado.xlsx')

planilha = load_workbook('C:\\Gerador Mega Sena\\bases\\sorteios.xlsx')

aba_ativa = planilha.active

aba_ativa.column_dimensions['I'].width = 26
aba_ativa['I1'] = 'Combinação (Pares/Ímpares)'

for celula in aba_ativa['I']:
    if celula.value != 'Combinação (Pares/Ímpares)':
        linha = celula.row
        qtde_par = 0
        qtde_impar = 0
        for col in range(3,9):
            coluna = get_column_letter(col)
            valor = aba_ativa[f'{coluna}{str(linha)}'].value
            if par_impar(valor) == True:
                qtde_par += 1
            elif par_impar(valor) == False:
                qtde_impar += 1
        aba_ativa[f'I{linha}'] = f'{qtde_par}{qtde_impar}'
    else:
        pass

aba_ativa.column_dimensions['K'].width = 12
aba_ativa['K1'] = 'Combinação'
aba_ativa.column_dimensions['L'].width = 10
aba_ativa['L1'] = 'Repetição'

numero = 0
for linha in range (2,9):
    aba_ativa[f'K{linha}'] = f'{numero}{6 - numero}'
    aba_ativa[f'L{linha}'] = Translator("=COUNTIF(I:I,K2)", origin="L2").translate_formula(f'L{linha}')
    numero += 1

comb1 = 0
comb2 = 0
comb3 = 0
comb4 = 0
comb5 = 0
comb6 = 0
comb7 = 0

for celula in aba_ativa['I']:
    if celula.value != 'Combinação (Pares/Ímpares)':
        linha = celula.row
        valor = aba_ativa[f'I{linha}'].value
        if valor == '06':
            comb1 += 1
        if valor == '15':
            comb2 += 1
        if valor == '24':
            comb3 += 1
        if valor == '33':
            comb4 += 1
        if valor == '42':
            comb5 += 1
        if valor == '51':
            comb6 += 1
        if valor == '60':
            comb7 += 1

combinacoes = []

combinacoes.append(comb1)
combinacoes.append(comb2)
combinacoes.append(comb3)
combinacoes.append(comb4)
combinacoes.append(comb5)
combinacoes.append(comb6)
combinacoes.append(comb7)

# numero = 0
# for i in range(0,7):    
#     print(f'A combinação de {numero} pares e {6 - numero} ímpares foi sorteada {combinacoes[i]} vezes')
#     numero += 1

maior_combinacao = combinacoes.index(max(combinacoes))

if maior_combinacao == 0: 
    num_par = 0
    num_impar = 6
if maior_combinacao == 1:
    num_par = 1
    num_impar = 5
if maior_combinacao == 2:
    num_par = 2
    num_impar = 4
if maior_combinacao == 3:
    num_par = 3
    num_impar = 3
if maior_combinacao == 4:
    num_par = 4
    num_impar = 2
if maior_combinacao == 5:
    num_par = 5
    num_impar = 1
if maior_combinacao == 6:
    num_par = 6
    num_impar = 0

planilha.save('C:\\Gerador Mega Sena\\bases\\sorteios_contabilizado.xlsx')